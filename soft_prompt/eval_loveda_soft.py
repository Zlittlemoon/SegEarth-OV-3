import os
import os.path as osp
import argparse
import openpyxl
from mmengine.runner import Runner
from mmengine.config import Config, DictAction

import segearthov3_segmentor_soft
import custom_datasets


def parse_args():
    parser = argparse.ArgumentParser(description='LoveDA soft-prompt evaluation with MMSeg')
    parser.add_argument('config', default='./configs/cfg_loveda_soft.py')
    parser.add_argument('--show', action='store_true')
    parser.add_argument('--show_dir', default='./show_dir/')
    parser.add_argument('--out', type=str, help='save raw predictions for offline evaluation')
    parser.add_argument(
        '--cfg-options',
        nargs='+',
        action=DictAction,
        help='override settings in config'
    )
    parser.add_argument(
        '--launcher',
        choices=['none', 'pytorch', 'slurm', 'mpi'],
        default='none'
    )
    parser.add_argument('--local_rank', '--local-rank', type=int, default=0)
    args = parser.parse_args()

    if 'LOCAL_RANK' not in os.environ:
        os.environ['LOCAL_RANK'] = str(args.local_rank)
    return args


def append_experiment_result(file_path, experiment_data):
    try:
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active
    if sheet['A1'].value is None:
        sheet['A1'] = 'Model'
        sheet['B1'] = 'Dataset'
        sheet['C1'] = 'aAcc'
        sheet['D1'] = 'mIoU'
        sheet['E1'] = 'mAcc'

    last_row = sheet.max_row
    for index, result in enumerate(experiment_data, start=1):
        sheet.cell(row=last_row + index, column=1, value=result.get('Model'))
        sheet.cell(row=last_row + index, column=2, value=result.get('Dataset'))
        sheet.cell(row=last_row + index, column=3, value=result.get('aAcc'))
        sheet.cell(row=last_row + index, column=4, value=result.get('mIoU'))
        sheet.cell(row=last_row + index, column=5, value=result.get('mAcc'))

    workbook.save(file_path)


def main():
    args = parse_args()
    cfg = Config.fromfile(args.config)
    cfg.launcher = args.launcher

    if args.out is not None:
        cfg.test_evaluator['output_dir'] = args.out
        cfg.test_evaluator['keep_results'] = True

    if args.cfg_options is not None:
        cfg.merge_from_dict(args.cfg_options)

    cfg.work_dir = osp.join('./work_dirs', osp.splitext(osp.basename(args.config))[0])

    runner = Runner.from_cfg(cfg)
    results = runner.test()

    model_name = cfg.model.get('type', 'UnknownModel')
    results.update({
        'Model': model_name,
        'Dataset': cfg.get('dataset_type', 'UnknownDataset')
    })

    if runner.rank == 0:
        append_experiment_result('results_loveda_soft.xlsx', [results])

        os.makedirs(cfg.work_dir, exist_ok=True)
        with open(os.path.join(cfg.work_dir, 'results.txt'), 'a') as f:
            f.write(os.path.basename(args.config).split('.')[0] + '\n')
            for k, v in results.items():
                f.write(f'{k}: {v}\n')


if __name__ == '__main__':
    main()