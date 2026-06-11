import os
import os.path as osp
import argparse
import openpyxl
from mmengine.runner import Runner
from mmengine.config import Config, DictAction

import segearthov3_segmentor_merge_sr_input
import custom_datasets


def parse_args():
    parser = argparse.ArgumentParser(description="SegEarthOV3 SAM3+SR-input evaluation")
    parser.add_argument("config", default="./configs/cfg_iSAID_sr_input.py")
    parser.add_argument("--show", action="store_true", help="show prediction results")
    parser.add_argument(
        "--show_dir",
        default="./show_dir/",
        help="directory to save visualization images",
    )
    parser.add_argument(
        "--out",
        type=str,
        help="The directory to save output prediction for offline evaluation",
    )
    parser.add_argument(
        "--cfg-options",
        nargs="+",
        action=DictAction,
        help="override config options, e.g. model.allow_bicubic_fallback=True",
    )
    parser.add_argument(
        "--launcher",
        choices=["none", "pytorch", "slurm", "mpi"],
        default="none",
        help="job launcher",
    )
    parser.add_argument("--local_rank", "--local-rank", type=int, default=0)
    args = parser.parse_args()
    if "LOCAL_RANK" not in os.environ:
        os.environ["LOCAL_RANK"] = str(args.local_rank)
    return args


def append_experiment_result(file_path, experiment_data):
    try:
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active
    if sheet["A1"].value is None:
        sheet["A1"] = "Model"
        sheet["B1"] = "Dataset"
        sheet["C1"] = "aAcc"
        sheet["D1"] = "mIoU"
        sheet["E1"] = "mAcc"

    last_row = sheet.max_row
    for index, result in enumerate(experiment_data, start=1):
        aAcc = result.get("aAcc", result.get("val/aAcc"))
        mIoU = result.get("mIoU", result.get("val/mIoU"))
        mAcc = result.get("mAcc", result.get("val/mAcc"))

        if aAcc is None or mIoU is None or mAcc is None:
            raise KeyError(f"Metric keys not found. Available keys: {list(result.keys())}")

        sheet.cell(row=last_row + index, column=1, value=result["Model"])
        sheet.cell(row=last_row + index, column=2, value=result["Dataset"])
        sheet.cell(row=last_row + index, column=3, value=aAcc)
        sheet.cell(row=last_row + index, column=4, value=mIoU)
        sheet.cell(row=last_row + index, column=5, value=mAcc)

    workbook.save(file_path)


def main():
    args = parse_args()
    print(os.getcwd())
    cfg = Config.fromfile(args.config)
    cfg.launcher = args.launcher

    if args.out is not None:
        cfg.test_evaluator["output_dir"] = args.out
        cfg.test_evaluator["keep_results"] = True
    if args.cfg_options is not None:
        cfg.merge_from_dict(args.cfg_options)

    cfg.work_dir = osp.join("./work_dirs", osp.splitext(osp.basename(args.config))[0])
    runner = Runner.from_cfg(cfg)
    results = runner.test()

    results.update({"Model": cfg.model.model_type, "Dataset": cfg.dataset_type})

    if runner.rank == 0:
        append_experiment_result("results_sam3_sr_input.xlsx", [results])
        with open(os.path.join(cfg.work_dir, "results_sam3_sr_input.txt"), "a") as f:
            f.write(os.path.basename(args.config).split(".")[0] + "\n")
            for key, value in results.items():
                f.write(key + ": " + str(value) + "\n")


if __name__ == "__main__":
    main()
